import xmlrpc.client

url = "http://odootst.fernuni.ch:8069"
db = "fsch_test"
username = "1533"
password = ""

import openpyxl
from openpyxl import Workbook

import argparse
import logging
import sys

_logger = logging.getLogger()
_logger.setLevel(logging.DEBUG)
ch = logging.StreamHandler()
ch.setLevel(logging.DEBUG)
formatter = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")
ch.setFormatter(formatter)
_logger.addHandler(ch)

class TranslationUnification(object):
    def __init__(self, sysArgs):
        
        
        self.args = argparse.ArgumentParser()
        self.args.add_argument(
            '-e',
            '--exportTranslations',
            action='store_true',
            help='Export mode'
        )
        self.args.add_argument(
            '-i',
            '--importTranslations',
            metavar="FILENAME",
            help='Import mode. Must be followed by the name of an Excel file.'
        )
        self.args.add_argument(
            '-uid',
            '--user-id',
            type=int,
            help='User ID representing the language'
        )
        self.args.add_argument(
            '-m',
            '--model',
            type=str,
            help='Model name, as example product.product'
        )
        self.args.add_argument(
            '-f',
            '--field',
            type=str,
            help='Field name, as example, description or name'
        )
        self.args = self.args.parse_args()

        # Handle connection
        self.url = "http://odootst.fernuni.ch:8069"
        self.db = "fsch_test"
        self.username = "1533"
        self.password = "T2F1+6heRFEi"

        # Step 1: Authenticate and get the user id
        common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url))
        self.uid = common.authenticate(db, username, password, {})

        # Step 2: Connect to the object that you want to manipulate
        self.models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url), allow_none=True)

        # Step 3: Fetch the list of active languages
        self.active_languages = self.models.execute_kw(self.db, self.uid, self.password, 'res.lang', 'search_read', [[]], {'fields': ['code', 'name']})

        if self.args.exportTranslations and (not self.args.model or not self.args.field):
            raise argparse.ArgumentError(None, "--model and --field are required when using -e/--exportTranslations")
        
        # Check if the importTranslations argument is provided without a file name
        #if self.args.importTranslations is None:
        #    raise argparse.ArgumentError(None, "Error: Please provide an Excel file name using the -i option")

    def export_translations(self):
        _logger.info("###### Exporting operation #######")
        wb = Workbook()
        ws = wb.active
        ws.title = "Translations"
        
        # Getting all the records for the model
        records = self.models.execute_kw(
            self.db, self.uid, self.password, self.args.model, 'search_read',
            [],
            {
                'fields': ['id', self.args.field],
                'context': {'lang': 'en_US'}  # Assuming English as a base language, you can adjust as needed
            }
        )

        if not records:
            _logger.error(f"No records found on model {self.args.model}")
            return
        
        # Preparing the header row
        header = ["Record ID"]
        for lang in self.active_languages:
            header.append(lang['code'])
        ws.append(header)
        
        for record in records:
            record_id = record['id']
            row_data = [record_id]
            
            for lang in self.active_languages:
                lang_code = lang['code']
                # Fetching field value for the specific language
                field_value = self.models.execute_kw(
                    self.db, self.uid, self.password, self.args.model, 'read',
                    [record_id],
                    {
                        'fields': [self.args.field],
                        'context': {'lang': lang_code}
                    }
                )[0][self.args.field]
                row_data.append(field_value)
            _logger.info(f"New row:{row_data}")
            ws.append(row_data)

        wb.save("translations.xlsx")

    def import_translations(self):
        _logger.info("###### Importing operation #######")
        
        wb = openpyxl.load_workbook("translations.xlsx")
        ws = wb.active

        # Assuming the first row contains the header
        header = [cell.value for cell in ws[1]]
        
        if "Record ID" not in header:
            _logger.error("Invalid Excel file. 'Record ID' column not found!")
            return

        for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header row
            record_id = row[0]
            translations = {header[i]: value for i, value in enumerate(row) if i != 0 and value}

            for lang_code, value in translations.items():
                if lang_code not in [lang['code'] for lang in self.active_languages]:
                    _logger.warning(f"Language '{lang_code}' is not active or doesn't exist in the system!")
                    continue
                # Updating the translation for this language
                
                print(f"Row information: {row}")
                print(f"Lang Code {lang_code}, value {value}")
                try:
                    self.models.execute_kw(
                        self.db, self.uid, self.password, self.args.model, 'write',
                        [record_id, {self.args.field: value}],
                        {'context': {'lang': lang_code}}
                    )
                    _logger.info(f"Updated record {record_id} for language {lang_code} with value: {value}")
                except Exception as e:
                    _logger.error(f"Error updating record {record_id} for language {lang_code}, field {self.args.field}, value {value}: {str(e)}")

        _logger.info("###### Importing operation completed #######")

if __name__ == '__main__':
    translationuni = TranslationUnification(sys.argv[1:])
    if translationuni.args.exportTranslations:
        translationuni.export_translations()
    if translationuni.args.importTranslations:
        translationuni.import_translations()
